# =====================================================================
# XSD → Excel automatizace
# ---------------------------------------------------------------------
# Tento skript načte XML Schema (soubor .xsd) a vytvoří přehled
# všech „leaf“ elementů.
#
# Pro každý element se na prvním listu Excelu uloží:
#  - jeho hierarchie (úrovně, cesta, ISO status)
#  - typ a z něj odvozený formát (např. X(35), ISODate, boolean apod.)
#
# Ve druhém listu Excelu se vytvoří seznam všech unikátních typů 
# a jejich odvozených formátů. První list pak používá funkci VLOOKUP 
# pro automatické doplnění formátu k elementům.
#
# =====================================================================

import re
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ===============================================================
# Načtení XSD souboru
# ===============================================================

while True:
    xsd_filename = input("Zadej název XSD souboru: ").strip()

    # Automaticky doplní .xsd, pokud uživatel zapomene
    if not xsd_filename.lower().endswith(".xsd"):
        xsd_filename += ".xsd"

    xsd_path = os.path.abspath(xsd_filename)

    if os.path.exists(xsd_path):
        break

    print(f"❌ Soubor '{xsd_filename}' nebyl nalezen. Zkus to znovu.")

# Výstupní Excel má stejný název jako XSD
output_excel = os.path.splitext(xsd_filename)[0] + ".xlsx"

# Načtení XML stromu
tree = ET.parse(xsd_path)
root = tree.getroot()


# ===============================================================
# Pomocná funkce pro odstranění XML namespace
# ===============================================================
def strip_ns(tag):
    """Vrátí název elementu bez namespace (např. {xs}element → element)."""
    return tag.split('}')[-1] if '}' in tag else tag


# ===============================================================
# Načtení všech simpleType definic a jejich restrikcí
# ---------------------------------------------------------------
# Ukládáme:
# - base typ (xs:string, xs:boolean, xs:decimal, ...)
# - délky
# - pattern
# - enumerace
#
# Výstup:
# simple_types[type_name] = {
#   base, minLength, maxLength, totalDigits, fractionDigits, pattern, enumeration
# }
# ===============================================================

simple_types = {}
for simple_type in root.findall(".//{*}simpleType"):
    type_name = simple_type.get("name")
    if not type_name:
        continue

    restriction = simple_type.find(".//{*}restriction")
    data = {
        "base": None,
        "minLength": None,
        "maxLength": None,
        "totalDigits": None,
        "fractionDigits": None,
        "pattern": None,
        "enumeration": []
    }

    if restriction is not None:
        data["base"] = restriction.get("base")
        for child in restriction:
            tag = strip_ns(child.tag)
            if tag == "minLength":
                data["minLength"] = child.get("value")
            elif tag == "maxLength":
                data["maxLength"] = child.get("value")
            elif tag == "totalDigits":
                data["totalDigits"] = child.get("value")
            elif tag == "fractionDigits":
                data["fractionDigits"] = child.get("value")
            elif tag == "pattern":
                data["pattern"] = child.get("value")
            elif tag == "enumeration":
                data["enumeration"].append(child.get("value"))

    simple_types[type_name] = data


# ===============================================================
# Načtení všech complexType a jejich vnořených elementů
# ---------------------------------------------------------------
# - klasické sequence / choice
# - simpleContent / extension (např. boolean wrappery)
#
# Výstup:
# types[type_name] = list(child_elements)
# complex_simple_base[type] = base_simple_type
# ===============================================================

types = {}
complex_simple_base = {} 

for complex_type in root.findall(".//{*}complexType"):
    type_name = complex_type.get("name")
    if not type_name:
        continue

    # --- simpleContent / extension ---
    simple_content = complex_type.find(".//{*}simpleContent")
    if simple_content is not None:
        ext = simple_content.find(".//{*}extension")
        if ext is not None:
            base = ext.get("base")
            if base:
                complex_simple_base[type_name] = base
        continue

    # --- klasické sequence / choice ---
    elems = []
    for container in complex_type.findall(".//{*}sequence") + complex_type.findall(".//{*}choice"):
        is_choice = (strip_ns(container.tag) == "choice")
        for el in container.findall("{*}element"):
            if el.get("name"):
                elems.append({
                    "name": el.get("name"),
                    "type": el.get("type"),
                    "minOccurs": el.get("minOccurs", "1"),
                    "maxOccurs": el.get("maxOccurs", "1"),
                    "is_choice": is_choice
                })

    if elems:
        types[type_name] = elems
# ===============================================================
# Detekce complexType s <xs:any namespace="##any">
# ---------------------------------------------------------------
# Tyto typy budou mít formát <ANY>
# ===============================================================

complex_any_types = set()

for complex_type in root.findall(".//{*}complexType"):
    type_name = complex_type.get("name")
    if not type_name:
        continue

    # xs:any namespace="##any" → any
    any_elem = complex_type.find(".//{*}any")
    if any_elem is not None:
        ns = any_elem.get("namespace")
        if ns == "##any":
            complex_any_types.add(type_name)
            continue

# ===============================================================
# Nalezení root elementu dokumentu
# ===============================================================

root_element = None
root_type = None
for el in root.findall("./{*}element"):
    if el.get("type") in types:
        root_element = el.get("name")
        root_type = el.get("type")
        break

if not root_element:
    raise ValueError("❌ Nelze najít root element.")


# ===============================================================
# Odhad maximální délky z regex patternu
# ---------------------------------------------------------------
# Heuristický výpočet:
# - skupiny ( ... ){m,n}
# - znakové třídy [ ... ]{m,n}
# - pevné znaky a escapované znaky
# ===============================================================

def estimate_length_from_pattern(pattern: str) -> int | None:
    """
    Vypočítá maximální délku hodnoty z XSD regex patternu.
    """

    if not pattern:
        return None

    total = 0
    p = pattern

    # Skupiny ( ... ){m,n} nebo ( ... ){n} rekurzivně
    group_regex = re.compile(r'\(([^()]+)\)\{(\d+)(?:,(\d+))?\}')
    while True:
        match = group_regex.search(p)
        if not match:
            break
        inner, min_q, max_q = match.groups()
        inner_len = estimate_length_from_pattern(inner) or 0
        repeat = int(max_q or min_q)
        total += inner_len * repeat
        # Odstraníme z patternu zpracovanou skupinu
        p = p[:match.start()] + p[match.end():]

    # Znakové třídy [ ... ] {m,n}
    charclass_regex = re.compile(r'\[([^\]]+)\](?:\{(\d+)(?:,(\d+))?\})?')
    while True:
        match = charclass_regex.search(p)
        if not match:
            break
        # délka = kvantifikátor * 1 znak
        min_q, max_q = match.group(2), match.group(3)
        repeat = int(max_q or min_q or 1)
        total += repeat
        # Odstraníme z patternu zpracovanou třídu
        p = p[:match.start()] + p[match.end():]

    # Zbývající pevné znaky (včetně escaped)
    # \+ = 1 znak, 4 = 1 znak, - = 1 znak
    char_with_quant = re.compile(r'(\\?.)(?:\{(\d+)(?:,(\d+))?\})?')
    for match in char_with_quant.finditer(p):
        char, min_q, max_q = match.groups()
        repeat = int(max_q or min_q or 1)
        total += repeat

    return total if total > 0 else None
 
# ===============================================================
# Detekce typu znaků v regex patternu
# ---------------------------------------------------------------
# Výstup:
# - digits   → pouze číslice
# - letters  → pouze písmena
# - alnum    → písmena + číslice
# - text     → obsahuje jiné znaky (+ - / atd.)
# ===============================================================
   
def pattern_char_types(pattern: str):
    """Vrací typ znaků z patternu (digits, letters, alnum)."""
    if not pattern:
        return None

    # Odstranění konstrukce, které NESMÍ ovlivňovat typ
    cleaned = pattern

    # odstraníme znakové třídy a skupiny
    cleaned = re.sub(r'\[[^\]]+\]', '', cleaned)
    cleaned = re.sub(r'\([^\)]+\)', '', cleaned)

    # odstraníme kvantifikátory {m,n}
    cleaned = re.sub(r'\{\d+(?:,\d+)?\}', '', cleaned)

    # Detekce typů – ALE z PŮVODNÍHO patternu
    has_digit = bool(re.search(r'\d', pattern))
    has_alpha = bool(re.search(r'[A-Za-z]', pattern))

    # Zbylé "pevné" znaky (např. + - / :)
    has_other = bool(re.search(r'[^A-Za-z0-9]', cleaned))

    if has_other:
        return 'text'
    if has_digit and not has_alpha:
        return 'digits'
    if has_digit and has_alpha:
        return 'alnum'
    if has_alpha:
        return 'letters'
    return None


# ===============================================================
# 🎯 Odvození formátu (ISO / X / XN / boolean / <ANY>)
# ---------------------------------------------------------------
# Priorita:
# 1) boolean
# 2) enumeration
# 3) pattern
# 4) decimal
# 5) maxLength
# 6) heuristika z názvu typu
# ===============================================================

def derive_format(type_name: str, restrictions: dict):
    if not restrictions:
        return ""

    enum = restrictions.get("enumeration") or []
    pattern = restrictions.get("pattern")
    max_len = restrictions.get("maxLength")
    total_dig = restrictions.get("totalDigits")
    frac_dig = restrictions.get("fractionDigits")
    base_type = restrictions.get("base")

# Boolean
    if base_type:
        base = base_type.lower()
        if base in ("xs:boolean", "boolean"):
            return "boolean"
# Enumerace
    if enum:
        return f"X({max(len(str(v)) for v in enum)})"

# Pattern
    if pattern:
        est = estimate_length_from_pattern(pattern)
        ctype = pattern_char_types(pattern)
        if est:
            return f"XN({est})" if ctype == "digits" else f"X({est})"
        return ""

# Decimal - dynamicky z restrikcí
    if total_dig:
        frac = frac_dig or "0"
        return f"N({total_dig},{frac})"

# Textové délky
    if max_len:
        return f"X({max_len})"

# Odvození z názvu
    special = {
        "ISODateTime": "ISODateTime",
        "ISOYearMonth": "ISOYearMonth",
        "ISODate": "ISODate",
        "ISOTime": "ISOTime",
        "SupplementaryDataEnvelope1": "<any>",
        "SkipPayload": "<any>",
        "LanguageCode": "X(2)",
    }

    if type_name:
        t = type_name.lower()

        for k in sorted(special, key=len, reverse=True):
            if k in type_name:
                return special[k]

        m = re.search(r"Max(\d+).*Text", type_name)
        if m:
            return f"X({m.group(1)})"

        m = re.search(r"Max(\d+).*Numeric", type_name)
        if m:
            return f"XN({m.group(1)})"

    return ""

# ===============================================================
# Rekurzivní průchod strukturou XSD
# ---------------------------------------------------------------
# - sestavení cesty elementu
# - výpočet ISO statusu (M/O/C)
# - sběr pouze leaf elementů
# ===============================================================

rows = []

def format_occurrence(min_occurs, max_occurs):
    """Vrací formátovaný řetězec [min...max] pro výskyty."""
    max_disp = "∞" if max_occurs == "unbounded" else max_occurs
    return f"[{min_occurs}...{max_disp}]"

def traverse(element_name, element_type, minOccurs="1", maxOccurs="1",
             parent_optional_found=False, in_choice=False, path=None):
    """Rekurzivně prochází strukturu typů a vytváří úplnou hierarchii."""
    if path is None:
        path = []

    children = types.get(element_type, [])
    show_occ = False
    if children:
        if maxOccurs == "unbounded" or (str(maxOccurs).isdigit() and int(maxOccurs) > 1):
            show_occ = True

    name_with_occ = f"{element_name}{format_occurrence(minOccurs, maxOccurs)}" if show_occ else element_name
    current_path = path + [name_with_occ]

    # ISO status podle minOccurs a choice
    if in_choice:
        iso_status = "C"
    elif str(minOccurs) == "0":
        iso_status = "O"
        parent_optional_found = True
    elif parent_optional_found:
        iso_status = "C"
    else:
        iso_status = "M"

# Leaf element
    if not children:
        if element_type in complex_any_types:
            fmt = "<ANY>"
            restrictions = {}
        else:
            base_type = complex_simple_base.get(element_type)
            if base_type:
                restrictions = simple_types.get(base_type, {})
                fmt = derive_format(base_type, restrictions)
            else:
                restrictions = simple_types.get(element_type, {})
                fmt = derive_format(element_type, restrictions)

        rows.append({
            "path": current_path,
            "iso_status": iso_status,
            "type_name": element_type,
            "pattern": restrictions.get("pattern", ""),
            "enumeration": ", ".join(restrictions.get("enumeration", [])) if restrictions else "",
            "format": fmt
        })
    else:
        for child in children:
            traverse(
                child["name"],
                child["type"],
                child["minOccurs"],
                child["maxOccurs"],
                parent_optional_found,
                in_choice or child.get("is_choice", False),
                current_path
            )


# Spusť průchod od hlavního elementu
traverse(root_element, root_type)


# ===============================================================
# Generování Excel souboru
# ===============================================================

wb = Workbook()
ws = wb.active
ws.title = "Hierarchy"
ws["A1"] = os.path.basename(xsd_path)

if not rows:
    raise ValueError("Žádné leaf elementy nenalezeny.")

max_depth = max(len(r["path"]) for r in rows)
headers = [f"Level {i}" for i in range(1, max_depth + 1)] + [
    "Full Path", "Type name", "ISO Status", "Format", "Patterns", "Enumerations"
]
ws.append(headers)

start_row = 3
for r in rows:
    row_levels = r["path"] + [""] * (max_depth - len(r["path"]))
    ws.append(row_levels + ["", r["type_name"], r["iso_status"], "", r["pattern"], r["enumeration"]])

# ---------- Vytvoření formule pro sloupec "Full Path" ----------
full_path_col = max_depth + 1
type_col = max_depth + 2
format_col = max_depth + 4

for i in range(len(rows)):
    excel_row = start_row + i
    parts = []
    for col_idx in range(1, max_depth + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            parts.append(f'IF({col_letter}{excel_row}="","",{col_letter}{excel_row})')
        else:
            parts.append(f'IF({col_letter}{excel_row}="","","/" & {col_letter}{excel_row})')
    ws.cell(row=excel_row, column=full_path_col).value = "=" + " & ".join(parts)

# ---------- Druhý list: přehled typů ----------
ws_types = wb.create_sheet("Types")
ws_types.append(["type", "format", "minLength", "maxLength", "totalDigits",
                 "fractionDigits", "pattern", "enumeration"])

used_types = sorted({r["type_name"] for r in rows if r["type_name"]})

for t in used_types:
    if t in complex_any_types:
        fmt = "<ANY>"
        st = {}
    else:
        base_type = complex_simple_base.get(t)
        if base_type:
            st = simple_types.get(base_type, {})
            fmt = derive_format(base_type, st)
        else:
            st = simple_types.get(t, {})
            fmt = derive_format(t, st)

    ws_types.append([
        t, fmt,
        st.get("minLength", ""),
        st.get("maxLength", ""),
        st.get("totalDigits", ""),
        st.get("fractionDigits", ""),
        st.get("pattern", ""),
        ", ".join(st.get("enumeration", [])) if st else ""
    ])

# ---------- VLOOKUP vzorce pro formát ----------
for i in range(len(rows)):
    excel_row = start_row + i
    type_cell = f"{get_column_letter(type_col)}{excel_row}"
    vlookup_expr = f'IFERROR(IF(VLOOKUP({type_cell},Types!$A:$H,2,FALSE)="","ERROR",VLOOKUP({type_cell},Types!$A:$H,2,FALSE)),"ERROR")'
    ws.cell(row=excel_row, column=format_col).value = f"={vlookup_expr}"


# ===============================================================
# Uložení výstupního Excelu
# ===============================================================

wb.save(output_excel)
print(f"✅ Hotovo! Vytvořen soubor '{output_excel}'.")
